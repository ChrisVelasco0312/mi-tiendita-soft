from textual.screen import Screen
from textual.containers import Container, Horizontal
from textual.widgets import Static, Button, Input
from textual.reactive import reactive
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook

EXCEL_PATH = os.path.join("src", "business", "data", "product_stock_data.xlsx")


class CreateSellView(Screen):
    codigo = reactive("")
    cantidad = reactive("")
    total = reactive("0")
    producto = reactive(None)

    def compose(self):
        self.codigo_input = Input(placeholder="Código del ítem", id="codigo")
        self.cantidad_input = Input(placeholder="Cantidad a vender", id="cantidad")
        self.total_display = Static("Total a pagar: $0.00", id="total")
        self.detalles_producto = Static("", id="detalles")
        self.mensaje_resultado = Static("", id="mensaje")

        yield Container(
            Horizontal(
                Button("Volver", id="volver"),
                Button("Realizar Venta", id="realizar_venta")
            ),
            Static("Generar Nueva Venta"),
            self.codigo_input,
            self.detalles_producto,
            self.cantidad_input,
            self.total_display,
            self.mensaje_resultado
        )

    def get_product_by_code_excel(self, item_code):
        if not os.path.exists(EXCEL_PATH):
            return None

        wb = load_workbook(EXCEL_PATH)

        if "ProductData" not in wb.sheetnames:
            return None

        ws = wb["ProductData"]

        item_code = str(item_code).strip().lower()

        for row in ws.iter_rows(min_row=2, values_only=False):
            cell_value = row[0].value
            if cell_value is None:
                continue
            if str(cell_value).strip().lower() == item_code:
                return {
                    "row": row,
                    "sheet": ws,
                    "wb": wb,
                    "id": cell_value,
                    "categoria": row[1].value,   # <-- Agregado
                    "nombre": row[2].value,
                    "cantidad": int(row[3].value),
                    "precio": float(row[5].value)
                }

        return None

    def on_input_submitted(self, event: Input.Submitted):
        if event.input.id == "codigo":
            pid = event.value.strip()
            self.codigo = pid
            producto = self.get_product_by_code_excel(pid)
            if producto:
                self.producto = producto
                self.detalles_producto.update(
                    f"Nombre: {producto['nombre']} | Categoría: {producto['categoria']} | Disponible: {producto['cantidad']} | Precio: ${producto['precio']:.2f}"
                )
            else:
                self.producto = None
                self.detalles_producto.update("Producto no encontrado.")

        elif event.input.id == "cantidad":
            if self.producto:
                try:
                    cant = int(event.value.strip())
                    total = cant * self.producto["precio"]
                    self.cantidad = str(cant)
                    self.total = f"{total:.2f}"
                    self.total_display.update(f"Total a pagar: ${self.total}")
                except ValueError:
                    self.total_display.update("Cantidad inválida.")

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "volver":
            self.app.pop_screen()

        elif event.button.id == "realizar_venta":
            if self.producto and self.cantidad.isdigit():
                cant_vendida = int(self.cantidad)

                if cant_vendida <= self.producto["cantidad"]:
                    nueva_cantidad = self.producto["cantidad"] - cant_vendida
                    fila = self.producto["row"]
                    fila[3].value = nueva_cantidad  # Actualiza stock
                    self.producto["wb"].save(EXCEL_PATH)

                    fecha_venta = datetime.now().strftime("%Y-%m-%d")

                    # Guardar en hoja SalesData
                    wb = load_workbook(EXCEL_PATH)
                    if "SalesData" in wb.sheetnames:
                        ws_sales = wb["SalesData"]
                    else:
                        ws_sales = wb.create_sheet("SalesData")

                    # Si es la primera fila, escribe encabezados incluyendo "Categoría"
                    if ws_sales.max_row == 1 and ws_sales["A1"].value is None:
                        ws_sales.append(["ID", "Categoría", "Nombre", "Cantidad Vendida", "Total", "Fecha de Venta"])

                    # Registrar datos de venta
                    ws_sales.append([
                        self.producto["id"],
                        self.producto["categoria"],
                        self.producto["nombre"],
                        cant_vendida,
                        float(self.total),
                        fecha_venta
                    ])
                    wb.save(EXCEL_PATH)

                    self.mensaje_resultado.update("Venta realizada exitosamente.")
                    self.detalles_producto.update("")
                    self.codigo_input.value = ""
                    self.cantidad_input.value = ""
                    self.total_display.update("Total a pagar: $0.00")
                    self.producto = None
                else:
                    self.mensaje_resultado.update("Cantidad mayor al stock disponible.")
            else:
                self.mensaje_resultado.update("Datos incompletos para realizar la venta.")