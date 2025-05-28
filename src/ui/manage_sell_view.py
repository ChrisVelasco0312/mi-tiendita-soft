from textual.screen import Screen
from textual.containers import Container, Horizontal
from textual.widgets import Static, Button, DataTable
from openpyxl import load_workbook
import os

EXCEL_PATH = os.path.join("src", "business", "data", "product_stock_data.xlsx")

class ManageSellView(Screen):

    def compose(self):
        tabla = DataTable(zebra_stripes=True)
        # Columnas solicitadas
        tabla.add_columns("Código", "Categoría", "Nombre", "Quantity", "Total", "date")

        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH)
            if "SalesData" in wb.sheetnames:
                ws = wb["SalesData"]

                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if all(cell is None for cell in row):
                        continue
                    try:
                        # Orden esperado: ID, Categoría, Nombre, Cantidad Vendida, Total, Fecha de Venta
                        codigo, categoria, nombre, quantity, total, date = row
                        tabla.add_row(
                            str(codigo),
                            str(categoria),
                            str(nombre),
                            str(quantity),
                            f"${total:.2f}",
                            str(date)
                        )
                    except Exception as e:
                        print(f"Error en fila {i}: {e}")

        yield Container(
            Static("CONSULTAR VENTAS", id="titulo"),
            Static("Ventas Realizadas", id="subtitulo"),
            tabla,
            Horizontal(
                Button("Volver", id="volver"),
                Button("Generar Venta", id="generar_venta")
            )
        )

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "volver":
            self.app.pop_screen()
        elif event.button.id == "generar_venta":
            # Navegar a la pantalla CreateSellView registrada como "create_sell"
            self.app.push_screen("create_sell_view")